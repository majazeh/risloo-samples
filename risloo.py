import unilities
import warnings
import json
import sys
from pathlib import Path



class risloo():
    def __init__(self, **args):
        self.args = args 
        self.my_dir = str(Path(__file__).resolve().parent)
        input_type = args['input_type']
        data = args['input_data']
        dash = ''
        if args['scale'][0].isdigit():
            dash = '_'
        # Scoring
        try:
            self.scoring_scale = getattr(__import__('scoring.' + dash + args['scale'] , fromlist=[args['scale']]), dash + args['scale'])(data, input_type)
            
        except:
            raise Exception('score scale `%s` not defined', args['scale'])
        self.scoring_scale.scoring()
    
        #Interpretation if it is needed
        if args['Interpretation']:
            try:
                self.interpreting_scale = getattr(__import__('interpreting.'  + args['scale'], fromlist=[args['scale']]), dash + args['scale'])(data, input_type)                
            except:
                raise Exception('interp scale `%s` not defined', args['scale'])
            
            self.interpreting_scale.interpreting()

    def export(self):
        for out_type in self.args['output_types']:            
            if out_type =='raw':
                self.export_raw()
            elif  out_type =='json':
                self.export_json()
            elif  out_type =='excell':
                self.export_excell()
  
    def export_raw(self):   
        sys.stdout.write(json.dumps(self.scoring_scale.score.toDict(),ensure_ascii=False) + "\n")       
        if self.args['Interpretation']:
            sys.stdout.write(self.interpreting_scale.interpret.get_text() + "\n")
    
    
    def export_excell(self):       
        import openpyxl
        import os
        from os.path import join as opj
        #excell
        my_path = opj(self.my_dir ,'scoring','tests')
        addr = self.args['input_data'].split('.json')[0]
        num = addr[-1]
        names = addr.split(os.path.sep)        
        test_name = names [-2]
        file_name = names [-1].split('_')[0]        
        excell_path = opj(my_path, test_name , names [-1] + '.xlsx')        
        Path(opj(my_path ,test_name)).mkdir(exist_ok=True)        
        book = openpyxl.load_workbook(excell_path )

        if "outputs"  in book.sheetnames:
             sheet = book["outputs"]
        else :
            book.create_sheet("outputs")
            sheet = book["outputs"]
   
        sheet.cell(row = 1, column=2).value = 'پاسخ سیستم'
        sheet.cell(row = 1, column=3).value = 'پاسخ دستی کارشناس'
        
        out_dic = self.scoring_scale.score.toDict()
        cnt = 2 
        for key in out_dic.keys():       
            sheet.cell(row= cnt  , column=1).value = key
            if isinstance(out_dic[key], dict):          
                
                for new_key in out_dic[key].keys():
                    cnt = cnt + 1         
                    sheet.cell(row= cnt , column=1).value = new_key
                    sheet.cell(row= cnt, column=2).value = out_dic[key][new_key]    
                cnt = cnt + 1 
            else :  
                sheet.cell(row= cnt, column=2).value = out_dic[key]
                cnt = cnt + 1
        book.save(excell_path)
    
    def export_json(self) :
        #json
        import os
        from os.path import join as opj
        my_path = opj(self.my_dir , 'scoring', 'jsons')
        addr = self.args['input_data'].split('.json')[0]
        names = addr.split(os.path.sep)  
        folder_name = names [-2]
        file_name = names [-1]
        Path(my_path + folder_name ).mkdir(exist_ok=True) 
        with open(opj(my_path, folder_name, file_name + '_result.json', 'w')) as file:
            json.dump(self.scoring_scale.score.toDict(), file, ensure_ascii=False ,indent= 4)
        
if (__name__ == '__main__'):
    rsl = risloo(**unilities.get_args())
    rsl.export()

