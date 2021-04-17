import json
from random import randint
from openpyxl import Workbook ,load_workbook
from pathlib import Path

my_path = str(Path(__file__).resolve().parent)
my_par_dir = my_path.split('/tests')[0]



########################### Parameter setting ################################
test_name = 'WAQ'
file_name = 'WAQ'
test_numbers = 2
test_numbers_list = [1,2]

# test form loading
with open( my_par_dir + '/forms/' + test_name + '/' + file_name +'.json', 'r') as file:
    data = json.load(file)


# Test producing
##################### user_answered #######################################
data["prerequisites"][0] ["user_answered"] = "2" # 1 for woman and 2 for man
data["prerequisites"][1] ["user_answered"] = "37" # age
data["prerequisites"][2] ["user_answered"] = "8"#str(randint(1,10)) # education options from 1 to 10

for j in range (test_numbers):
    book = Workbook()
    sheet = book.active
    
    sheet.title = "question_answers"
   
    sheet.cell(row = 1, column=1).value = 'شماره سوالات'
    sheet.cell(row = 1, column=2).value = 'پاسخ  گزینه فرد'
    
    
    for i,item in enumerate(data["items"]):
        num_options = len(item["answer"]["options"])
        item ["user_answered"] = str(randint(1,num_options))
        sheet.cell(row = i+2, column=1).value = i+1
        sheet.cell(row = i+2, column=2).value = item ["user_answered"]
    
        



    book.save( my_path + '/' + test_name +'/' + file_name + '_test_' + str(test_numbers_list[j]) + '.xlsx') 

    with open( my_path + '/' + test_name + '/' + file_name +'_test_'+ str(test_numbers_list[j]) + '.json', 'w') as file:
        json.dump(data, file , ensure_ascii=False ,indent= 4)


