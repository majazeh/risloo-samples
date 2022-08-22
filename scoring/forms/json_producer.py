
import json
from openpyxl import Workbook ,load_workbook
from pathlib import Path
from os.path import join as opj
import os
import argparse


## parse arguments
parser = argparse.ArgumentParser()
parser.add_argument("--folder_name" , required=True)
parser.add_argument("--file_name", required=True)
parser.add_argument("--title", required=True, nargs='*')
parser.add_argument("--description", required=True, nargs='*')
parser.add_argument("--version", default="")
parser.add_argument("--edition", default="")
parser.add_argument("--edition_version", default="")
parser.add_argument("--filler" , default="")
parser.add_argument("--option_positions_is_typical_form" , action="store_true", required=True)
parser.add_argument("--option_list_array", nargs='*')

args = parser.parse_args().__dict__



## map arguments to local variables
folder_name = args['folder_name']
file_name = args['file_name']
title = args['title']
description = args['description']
version = args['version']
edition = args['edition']
edition_version = args['edition_version']
filler = args['filler']
option_positions_is_typical_form = args['option_positions_is_typical_form']
options_list = args["option_list_array"]




my_path = str(Path(__file__).resolve().parent)
my_par_dir = my_path.split(os.path.sep +'forms')[0]


# raw form loading
with open( opj(my_path , 'raw.json'), 'r' ) as file:
    data = json.load(file)


########################### Parameter setting ################################

data["title"] = " ".join(title)

data["description"] = " ".join(description)

# data["version"] = ""
# data["edition"] = ""
# data["edition_version"]=
# data["filler"]=



###############################################################################
items = []
item ={}
item = data["items"][0].copy()


########################## load questions excell ###########################
print(opj(my_par_dir, 'tests', folder_name, file_name + '.xlsx'))
book = load_workbook(opj(my_par_dir, 'tests', folder_name, file_name + '.xlsx'))
sheet = book.active



question_numbers =  sheet.max_row - 1
print("question_numbers: ",question_numbers)


for i in range(question_numbers):
    item["text"] = str(sheet.cell(row = i+2, column=1).value)
    items.append(item.copy())

data["items"] = items




### Option_lists ####

if option_positions_is_typical_form:
    
    num_options = sheet.max_column - 1
    options_list = []
    
    for i in range(num_options):
        options_list.append(str(sheet.cell(row = 1, column=i+2).value))
    
        
else:

    num_options = len(options_list)

item["answer"]["options"] = options_list


###### make directory and dump results ###

Path(my_path + os.path.sep + folder_name ).mkdir(exist_ok=True)

with open(opj(my_path, folder_name, file_name + '.json'), 'w') as file:
    json.dump(data, file, ensure_ascii=False ,indent= 4)




