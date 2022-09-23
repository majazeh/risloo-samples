import json
import random 
from random import randint
from pathlib import Path
import subprocess
import argparse
import os
from os.path import join as opj


## parse arguments
parser = argparse.ArgumentParser()
parser.add_argument("--folder_name")
parser.add_argument("--file_name")
parser.add_argument("--test_numbers")
parser.add_argument("-s" , action="store_true")

args = parser.parse_args().__dict__

## map arguments to local variables
folder_name = args['folder_name']
file_name = args['file_name']
test_numbers = int(args['test_numbers'])
SAVE = args['s']

if SAVE :
    from openpyxl import Workbook ,load_workbook

test_numbers_list = list(range(1,test_numbers+1))

my_path = str(Path(__file__).resolve().parent)

# test form loading
with open(opj(my_path,'scoring', 'forms', folder_name, file_name +'.json'), 'r') as file:
    data = json.load(file)


for j in range (test_numbers):
    
    # Test producing
    ##################### user_answered #######################################
    data["prerequisites"][0] ["user_answered"] = "2" # 1 for woman and 2 for man
    data["prerequisites"][1] ["user_answered"] = "37" # age
    data["prerequisites"][2] ["user_answered"] = "8"#str(randint(1,10)) # education options from 1 to 10
    
    if SAVE :
        book = Workbook()
        sheet = book.active        
        sheet.title = "question_answers"
        sheet.cell(row = 1, column=1).value = 'شماره سوالات'
        sheet.cell(row = 1, column=2).value = 'پاسخ  گزینه فرد'    

    for i,item in enumerate(data["items"]):
        
        ## optional options
        if item["answer"]["type"] == "optional" :
            
            num_options = len(item["answer"]["options"])
            item ["user_answered"] = str(randint(1,num_options))
        
        
        ## range options
        elif item["answer"]["type"] == "range":
            
            if "reverse" in item["answer"]:
                if item["answer"]["reverse"]:
                    item ["user_answered"] = str(item["answer"]["max"] +1 - randint(1, item["answer"]["max"] - item["answer"]["min"] +1))
                
            else:
                item ["user_answered"] = str(randint(1, item["answer"]["max"] - item["answer"]["min"] +1))
        
        
        ## descriptive options
        elif item["answer"]["type"] == "descriptive":
            
            if random.random()>=0.5:
                item ["user_answered"] = "fake text"
            
            else:
                item ["user_answered"] = ""
         
         
            
        ## sortable options
        elif item["answer"]["type"] =='sortable' :
            
            num_options = len(item["answer"]["options"])
            random_options = list(range(1 ,num_options +1))
            random.shuffle(random_options) 
            
            for k in range(num_options):
                random_options[k] = str(random_options[k])
            
            item ["user_answered"] = ",".join(random_options)
            
        
          
        ## number options
        elif item["answer"]["type"] == 'number' :
            
            item ["user_answered"] = randint(1,100)
            
            
        ### matrix radio options    
        elif item["answer"]["type"]  == 'matrix_radio':
            
            num_horizontal_options = len(item["answer"]["matrix"][0])
            num_vertical_options = len(item["answer"]["matrix"][1]) 
            
            item ["user_answered"] = "({},{})".format(randint(1,num_horizontal_options) ,randint(1,num_vertical_options))
        

        if SAVE:
            sheet.cell(row = i+2, column=1).value = i+1
            print(item ["user_answered"])
            sheet.cell(row = i+2, column=2).value = item ["user_answered"]


    if SAVE:
        Path(opj(my_path, 'scoring', 'tests', folder_name)).mkdir(exist_ok=True) 

        book.save(opj(my_path, 'scoring', 'tests', folder_name, file_name + '_test_' + str(test_numbers_list[j]) + '.xlsx')) 

        with open(opj(my_path, 'scoring', 'tests', folder_name, file_name + '_test_' + str(test_numbers_list[j]) + '.json'), 'w') as file:
            json.dump(data, file , ensure_ascii=False ,indent= 4)

    if not SAVE:
        y = json.dumps(data,separators=(',', ':'))
        
        command = ["python3", "risloo.py" ,"-s",file_name ,"-it","raw","-id" ,y]       
        rc = subprocess.run(command,capture_output=True)
        print("test %d :" % (j+1) , rc.stdout.decode() + '\n\n')
    
    



